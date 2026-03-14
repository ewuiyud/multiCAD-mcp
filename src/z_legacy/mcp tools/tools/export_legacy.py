"""
Data export tools for extracting and saving drawing information.

Provides tools for:
- Extracting drawing data (entities with properties)
- Exporting data to Excel format
"""

import logging
import json
from typing import Optional

from mcp.server.fastmcp import Context

from mcp_tools.decorators import cad_tool, cad_tool_with_ui, get_current_adapter
from mcp_tools.helpers import result_message

logger = logging.getLogger(__name__)


def set_cell_value_safe(ws, row: int, col: int, value):
    """
    Write value to (row, col). If that address is inside a merged range,
    write to the merged range's top-left anchor instead.

    Args:
        ws: openpyxl worksheet
        row: Row number (1-indexed)
        col: Column number (1-indexed)
        value: Value to write
    """
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row=row, column=col)

    # Check if this cell is within a merged range
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                # Write to the anchor cell (top-left of merged range)
                ws.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                ).value = value
                return

    # Not a merged cell, write directly
    cell.value = value


def register_export_tools(mcp):
    """Register export tools with FastMCP.

    Args:
        mcp: FastMCP instance
    """

    @cad_tool(mcp, "export_drawing_to_excel")
    def export_drawing_to_excel(
        ctx: Context,
        filename: str = "drawing_data.xlsx",
        cad_type: Optional[str] = None,
    ) -> str:
        """Export drawing data to Excel file.

        Extracts all entities from the drawing and creates an Excel spreadsheet with:
        - Handle: Entity identifier
        - ObjectType: Entity type (LINE, CIRCLE, LWPOLYLINE, etc.)
        - Layer: Layer name
        - Color: Color index or name
        - Length: Length (for linear objects)
        - Area: Area (for closed objects)
        - Name: Name (for blocks, etc.)

        Files are saved to the output directory configured in config.json for security.
        If subdirectories are specified in the filename, they must be within the
        configured output directory.

        Args:
            filename: Excel filename or path (default: "drawing_data.xlsx")
                     Examples:
                     - "data.xlsx" → saved to output directory
                     - "exports/data.xlsx" → saved to output/exports/
            cad_type: CAD application to use (autocad, zwcad, gcad, bricscad)

        Returns:
            JSON result with export status
        """
        try:
            adapter = get_current_adapter()
            success = adapter.export_to_excel(filename)
            return result_message(
                "export drawing to excel",
                success,
                f"Saved to {filename}" if success else "Check logs for details",
            )
        except Exception as e:
            logger.error(f"Export failed: {e}")
            return result_message("export drawing to excel", False, str(e))

    @cad_tool_with_ui(mcp, "extract_drawing_data", ui_resource="drawing_viewer")
    def extract_drawing_data(
        ctx: Context,
        cad_type: Optional[str] = None,
    ) -> str:
        """Extract all drawing data without saving to file.

        Returns entity data as JSON with columns:
        - Handle: Entity identifier
        - ObjectType: Entity type (LINE, CIRCLE, LWPOLYLINE, etc.)
        - Layer: Layer name
        - Color: Color index or name
        - Length: Length (for linear objects)
        - Area: Area (for closed objects)
        - Name: Name (for blocks, etc.)

        In MCP Apps-compatible hosts (Claude Desktop, VS Code), this tool provides
        an interactive UI with filtering, sorting, and statistics visualization.

        Args:
            cad_type: CAD application to use (autocad, zwcad, gcad, bricscad)

        Returns:
            JSON result with extracted data and UI metadata
        """
        try:
            adapter = get_current_adapter()
            data = adapter.extract_drawing_data()

            if data:
                # Build result with UI resource
                result = {
                    "success": True,
                    "count": len(data),
                    "message": f"Extracted data from {len(data)} entities",
                    "entities": data,
                    "_meta": {
                        "ui": {
                            "resourceUri": "ui://multicad/drawing_viewer",
                            "data": {"entities": data},
                        }
                    },
                }
            else:
                result = {
                    "success": False,
                    "count": 0,
                    "message": "No entities found in drawing",
                    "entities": [],
                }

            return json.dumps(result, indent=2)

        except Exception as e:
            logger.error(f"Data extraction failed: {e}")
            result = {
                "success": False,
                "count": 0,
                "message": f"Extraction error: {str(e)}",
                "entities": [],
            }
            return json.dumps(result, indent=2)

    @cad_tool(mcp, "export_selected_to_excel")
    def export_selected_to_excel(
        ctx: Context,
        filename: str = "selected_data.xlsx",
        cad_type: Optional[str] = None,
    ) -> str:
        """Export only selected entities to Excel file.

        Exports only the currently selected entities to an Excel spreadsheet.
        If no entities are selected, returns an error message.

        Includes the same data columns as export_drawing_to_excel:
        - Handle: Entity identifier
        - ObjectType: Entity type
        - Layer: Layer name
        - Color: Color index or name
        - Length: Length (for linear objects)
        - Area: Area (for closed objects)
        - Name: Name (for blocks, etc.)

        Args:
            filename: Excel filename or path (default: "selected_data.xlsx")
            cad_type: CAD application to use (autocad, zwcad, gcad, bricscad)

        Returns:
            JSON result with export status
        """
        try:
            adapter = get_current_adapter()

            # Check if entities are selected
            if not adapter.has_selection():
                return result_message(
                    "export selected to excel",
                    False,
                    "No entities selected. Please select entities first.",
                )

            # Extract data for selected entities only
            from pathlib import Path
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from core.config import get_config

            config = get_config()

            # SECURITY: Resolve output directory first
            output_dir = Path(config.output.directory).expanduser().resolve()

            # Determine directory and filename
            if filename:
                dir_part = str(Path(filename).parent)
                if dir_part and dir_part != ".":
                    export_dir = dir_part
                else:
                    export_dir = str(output_dir)
            else:
                export_dir = str(output_dir)

            export_dir_path = Path(export_dir).expanduser().resolve()

            # SECURITY: Verify the directory is within the configured output directory
            try:
                export_dir_path.relative_to(output_dir)
            except ValueError:
                logger.error(
                    f"Security: Attempted to export outside output directory. "
                    f"Requested: {export_dir_path}, Allowed: {output_dir}"
                )
                return result_message(
                    "export selected to excel", False, "Invalid output directory"
                )

            # Create directory if it doesn't exist
            export_dir_path.mkdir(parents=True, exist_ok=True)

            # Get filename and construct full path
            file_obj = Path(filename)
            file_name = file_obj.name if filename else "selected_data.xlsx"
            full_filepath = export_dir_path / file_name

            # Extract selected data
            data = adapter.extract_drawing_data(only_selected=True)

            if not data:
                return result_message(
                    "export selected to excel",
                    False,
                    "Selected entities contain no exportable data",
                )

            # Create workbook
            workbook = Workbook()
            worksheet = workbook.active

            if worksheet is None:
                return result_message(
                    "export selected to excel",
                    False,
                    "Failed to create worksheet",
                )

            worksheet.title = "Selected Data"

            # Define columns
            columns = [
                "Handle",
                "ObjectType",
                "Layer",
                "Color",
                "Length",
                "Area",
                "Radius",
                "Circumference",
                "Name",
            ]

            # Write headers with styling
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")

            for col_idx, column_name in enumerate(columns, 1):
                try:
                    set_cell_value_safe(worksheet, 1, col_idx, column_name)
                    cell = worksheet.cell(row=1, column=col_idx)
                    if cell is not None:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                except Exception as e:
                    logger.debug(f"Error writing header at column {col_idx}: {e}")

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, column_name in enumerate(columns, 1):
                    try:
                        value = row_data.get(column_name)
                        set_cell_value_safe(worksheet, row_idx, col_idx, value)
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell is not None:
                            cell.alignment = Alignment(
                                horizontal="left", vertical="center"
                            )

                            # Apply number format for numeric columns
                            if column_name in [
                                "Length",
                                "Area",
                                "Radius",
                                "Circumference",
                            ] and isinstance(value, (int, float)):
                                cell.number_format = "0.000"
                    except Exception as e:
                        logger.debug(
                            f"Error writing data at row {row_idx}, col {col_idx}: {e}"
                        )

            # Auto-adjust column widths
            for col_idx, column_name in enumerate(columns, 1):
                max_length = len(column_name)
                for row_idx in range(2, len(data) + 2):
                    cell_obj = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = (
                        str(cell_obj.value or "") if cell_obj is not None else ""
                    )
                    max_length = max(max_length, len(cell_value))
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # Freeze first row
            worksheet.freeze_panes = "A2"

            # Save workbook
            workbook.save(str(full_filepath))
            logger.info(f"Exported {len(data)} selected entities to {full_filepath}")

            return result_message(
                "export selected to excel",
                True,
                f"Exported {len(data)} selected entities to {filename}",
            )

        except Exception as e:
            logger.error(f"Export selected failed: {e}")
            return result_message("export selected to excel", False, str(e))

    @cad_tool(mcp, "extract_selected_data")
    def extract_selected_data(
        ctx: Context,
        cad_type: Optional[str] = None,
    ) -> str:
        """Extract only selected entities data without saving to file.

        Returns entity data for selected entities as JSON with columns:
        - Handle: Entity identifier
        - ObjectType: Entity type
        - Layer: Layer name
        - Color: Color index or name
        - Length: Length (for linear objects)
        - Area: Area (for closed objects)
        - Name: Name (for blocks, etc.)

        If no entities are selected, returns an empty list.

        Args:
            cad_type: CAD application to use (autocad, zwcad, gcad, bricscad)

        Returns:
            JSON result with extracted selected data
        """
        try:
            adapter = get_current_adapter()

            # Check selection
            selection_info = adapter.get_selection_info()

            if selection_info["count"] == 0:
                result = {
                    "success": True,
                    "count": 0,
                    "message": "No entities selected",
                    "selected_count": 0,
                    "entities": [],
                }
                return json.dumps(result, indent=2)

            # Extract selected data
            data = adapter.extract_drawing_data(only_selected=True)

            if data:
                result = {
                    "success": True,
                    "count": len(data),
                    "message": f"Extracted data from {len(data)} selected entities",
                    "selected_count": selection_info["count"],
                    "entities": data,
                }
            else:
                result = {
                    "success": False,
                    "count": 0,
                    "message": "No exportable data in selected entities",
                    "selected_count": selection_info["count"],
                    "entities": [],
                }

            return json.dumps(result, indent=2)

        except Exception as e:
            logger.error(f"Selected data extraction failed: {e}")
            result = {
                "success": False,
                "count": 0,
                "message": f"Extraction error: {str(e)}",
                "selected_count": 0,
                "entities": [],
            }
            return json.dumps(result, indent=2)
