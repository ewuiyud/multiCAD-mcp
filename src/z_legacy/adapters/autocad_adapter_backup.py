"""
AutoCAD adapter for multiCAD-MCP.

Implements CADInterface for AutoCAD using Windows COM.
Supports AutoCAD, ZWCAD, GstarCAD, and BricsCAD via factory pattern.
"""

import logging
import time
import math
from typing import Optional, Dict, Any, List, Callable, TypeVar
from functools import wraps
from pathlib import Path
from contextlib import contextmanager
import sys

if sys.platform == "win32":
    import win32com.client
    import pythoncom
    import win32gui
    import win32api
    import win32con
    import pywintypes
else:
    raise ImportError("AutoCAD adapter requires Windows OS with COM support")

from core import (
    CADInterface,
    CADConnectionError,
    CADOperationError,
    InvalidParameterError,
    LayerError,
    ColorError,
    get_cad_config,
    get_config,
    Point,
    Coordinate,
)
from mcp_tools.constants import (
    COLOR_MAP,
    SS_COLOR_SELECT,
    SS_LAYER_SELECT,
    SS_TYPE_SELECT,
    SS_COPY,
    SS_SELECTION_GET,
    AUTOCAD_WINDOW_CLASSES,
    SELECTION_SET_IMPLIED,
    CLIPBOARD_DELAY,
    CLIPBOARD_STABILITY_DELAY,
    CLICK_DELAY,
    CLICK_HOLD_DELAY,
)

logger = logging.getLogger(__name__)

T = TypeVar("T")


# ========== COM Context Manager ==========


@contextmanager
def com_session():
    """Context manager for safe COM initialization and cleanup.

    Ensures CoInitialize/CoUninitialize are always paired, even on exceptions.
    Use this for all COM operations to prevent thread state leaks.

    Example:
        with com_session():
            app = win32com.client.Dispatch("AutoCAD.Application")
            # ... use app ...
    """
    pythoncom.CoInitialize()
    try:
        yield
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception as e:
            logger.debug(f"CoUninitialize failed (non-critical): {e}")


class SelectionSetManager:
    """Context manager for safe SelectionSet handling.

    Ensures SelectionSet cleanup even on exceptions, preventing orphaned
    selection sets that can cause issues in AutoCAD.

    Example:
        with SelectionSetManager(document, "TEMP_SS") as ss:
            ss.Select(...)
            # ... use ss ...
        # Auto-deleted on exit
    """

    def __init__(self, document: Any, name: str):
        """Initialize SelectionSet manager.

        Args:
            document: AutoCAD document object
            name: Name for the selection set
        """
        self.document = document
        self.name = name
        self.selection_set: Optional[Any] = None

    def __enter__(self) -> Any:
        """Create SelectionSet, deleting existing one if present.

        Returns:
            Created SelectionSet object
        """
        # Delete if exists
        try:
            self.document.SelectionSets.Item(self.name).Delete()
            logger.debug(f"Deleted existing SelectionSet: {self.name}")
        except Exception:
            pass

        # Create new
        self.selection_set = self.document.SelectionSets.Add(self.name)
        logger.debug(f"Created SelectionSet: {self.name}")
        return self.selection_set

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        """Cleanup SelectionSet on exit.

        Args:
            exc_type: Exception type if raised
            exc_val: Exception value if raised
            exc_tb: Exception traceback if raised
        """
        try:
            if self.selection_set:
                self.selection_set.Delete()
                logger.debug(f"Cleaned up SelectionSet: {self.name}")
        except Exception as e:
            logger.debug(f"Failed to delete SelectionSet {self.name}: {e}")


# ========== Decorators ==========


def com_safe(return_type: type = bool, operation_name: str = "operation"):
    """Decorator for COM operation error handling.

    Wraps method with:
    - Exception catching (pywintypes.com_error)
    - Operation logging
    - Automatic error conversion to CADOperationError

    Args:
        return_type: Expected return type (for type hints)
        operation_name: Name of operation (for logging)
    """

    def decorator(func: Callable[..., T]) -> Callable[..., T]:
        @wraps(func)
        def wrapper(*args, **kwargs) -> T:
            try:
                return func(*args, **kwargs)
            except pywintypes.com_error as e:
                # COM error attributes: args[0] = hresult, args[2] = strerror
                error_msg = f"COM error: {str(e)}"
                logger.error(f"Failed in {func.__name__}: {error_msg}")
                if return_type == bool:
                    return False  # type: ignore
                raise CADOperationError(operation_name, error_msg)
            except Exception as e:
                logger.error(f"Failed in {func.__name__}: {e}")
                if return_type == bool:
                    return False  # type: ignore
                raise CADOperationError(operation_name, str(e))

        return wrapper

    return decorator


# ========== AutoCAD Adapter ==========


class AutoCADAdapter(CADInterface):
    """Adapter for controlling AutoCAD via COM interface.

    Features:
    - Multi-CAD support (AutoCAD, ZWCAD, GstarCAD, BricsCAD) via cad_type parameter
    - Full drawing operations (lines, circles, arcs, polylines, dimensions, etc.)
    - Layer management (create, rename, delete, visibility control)
    - File operations (save, open, close, switch)
    - Entity selection and manipulation (move, rotate, scale, copy, paste)
    - Undo/redo support
    - Robust error handling with specific exception types
    """

    def __init__(self, cad_type: str = "autocad"):
        """Initialize AutoCAD adapter.

        Args:
            cad_type: Type of CAD (autocad, zwcad, gcad, bricscad)
        """
        self.cad_type = cad_type.lower()
        self.config = get_cad_config(self.cad_type)
        self.application = None
        self.document = None
        self._drawing_state: Dict[str, Any] = {
            "entities": [],
            "current_layer": "0",
        }

    # ========== Connection Management ==========

    def connect(self) -> bool:
        """Connect to CAD application with COM initialization."""
        try:
            logger.info(f"Connecting to {self.cad_type}...")

            # Initialize COM for this thread
            pythoncom.CoInitialize()

            # Try to get existing instance
            try:
                self.application = win32com.client.GetActiveObject(self.config.prog_id)
                logger.info(f"{self.cad_type} instance found (active)")
            except Exception:
                # Start new instance
                logger.info(f"{self.cad_type} not running, starting new instance...")
                try:
                    self.application = win32com.client.Dispatch(self.config.prog_id)
                except pywintypes.com_error as com_err:
                    error_code = com_err.args[0] if com_err.args else None
                    if error_code == -2147221005:
                        error_msg = (
                            f"Invalid ProgID '{self.config.prog_id}'. "
                            f"Either {self.cad_type.upper()} is not installed or the ProgID is incorrect. "
                            f"Check config.json and ensure the application is installed."
                        )
                    else:
                        error_msg = str(com_err)
                    logger.error(
                        f"Failed to create {self.cad_type} instance: {error_msg}"
                    )
                    raise CADConnectionError(self.cad_type, error_msg)

                if self.application is not None:
                    # Try to make application visible (not all CAD types support this)
                    try:
                        self.application.Visible = True
                    except (pywintypes.com_error, AttributeError) as e:
                        logger.debug(
                            f"{self.cad_type} doesn't support Visible property or it's read-only: {e}"
                        )
                self._wait_for(
                    lambda: self.application is not None,
                    timeout=self.config.startup_wait_time,
                )
                logger.info(
                    f"New {self.cad_type} instance started "
                    f"(waited {self.config.startup_wait_time}s)"
                )

            # Get active document or create new
            if self.application is not None:
                # Standard AutoCAD/ZWCAD/BricsCAD/GstarCAD handling
                if self.application.Documents.Count > 0:
                    self.document = self.application.ActiveDocument
                    logger.info("Using existing active document")
                else:
                    self.document = self.application.Documents.Add()
                    logger.info("Created new document")

            # Validate connection
            if not self._validate_document():
                raise CADConnectionError(self.cad_type, "Document validation failed")

            logger.info(f"✓ Successfully connected to {self.cad_type}")
            return True

        except pywintypes.com_error as e:
            error_msg = f"COM error: {str(e)}"
            logger.error(f"Failed to connect to {self.cad_type}: {error_msg}")
            raise CADConnectionError(self.cad_type, error_msg)
        except Exception as e:
            logger.error(f"Failed to connect to {self.cad_type}: {e}")
            raise CADConnectionError(self.cad_type, str(e))

    def disconnect(self) -> bool:
        """Disconnect from CAD application with COM cleanup."""
        try:
            if self.application:
                self.application = None
                self.document = None
            pythoncom.CoUninitialize()
            logger.info(f"Disconnected from {self.cad_type}")
            return True
        except Exception as e:
            logger.error(f"Error disconnecting: {e}")
            return False

    def __enter__(self):
        """Enter context: connect to CAD application.

        Allows using the adapter as a context manager:
            with AutoCADAdapter("autocad") as adapter:
                adapter.draw_line((0,0), (10,10))
                # Auto-disconnect on exit

        Returns:
            Self (the adapter instance)

        Raises:
            CADConnectionError: If connection fails
        """
        if not self.connect():
            raise CADConnectionError(
                self.cad_type, "Connection failed during context manager initialization"
            )
        return self

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        """Exit context: disconnect from CAD application.

        Args:
            exc_type: Exception type if raised
            exc_val: Exception value if raised
            exc_tb: Exception traceback if raised
        """
        self.disconnect()

    def is_connected(self) -> bool:
        """Check if connected to CAD application."""
        try:
            return (
                self.application is not None
                and self.document is not None
                and self._validate_document()
            )
        except Exception:
            return False

    def _validate_document(self) -> bool:
        """Validate that document is accessible."""
        try:
            if self.document is None:
                return False
            _ = self.document.Name
            return True
        except Exception:
            return False

    # ========== Drawing Operations ==========

    def draw_line(
        self,
        start: Coordinate,
        end: Coordinate,
        layer: str = "0",
        color: str | int = "white",
        lineweight: int = 0,
        _skip_refresh: bool = False,
    ) -> str:
        """Draw a line.

        Args:
            _skip_refresh: Internal flag to skip view refresh (used for batch operations)
        """
        document = self._get_document("draw_line")

        start_pt = CADInterface.normalize_coordinate(start)
        end_pt = CADInterface.normalize_coordinate(end)

        start_array = self._to_variant_array(start_pt)
        end_array = self._to_variant_array(end_pt)

        line = document.ModelSpace.AddLine(start_array, end_array)
        self._apply_properties(line, layer, color, lineweight)
        self._track_entity(line, "line")
        if not _skip_refresh:
            self.refresh_view()

        logger.debug(f"Drew line from {start} to {end}")
        return str(line.Handle)

    def draw_circle(
        self,
        center: Coordinate,
        radius: float,
        layer: str = "0",
        color: str | int = "white",
        lineweight: int = 0,
        _skip_refresh: bool = False,
    ) -> str:
        """Draw a circle.

        Args:
            _skip_refresh: Internal flag to skip view refresh (used for batch operations)
        """
        document = self._get_document("draw_circle")

        if radius <= 0:
            raise InvalidParameterError("radius", radius, "positive number")

        center_pt = CADInterface.normalize_coordinate(center)
        center_array = self._to_variant_array(center_pt)

        circle = document.ModelSpace.AddCircle(center_array, radius)
        self._apply_properties(circle, layer, color, lineweight)
        self._track_entity(circle, "circle")
        if not _skip_refresh:
            self.refresh_view()

        logger.debug(f"Drew circle at {center} with radius {radius}")
        return str(circle.Handle)

    def draw_arc(
        self,
        center: Coordinate,
        radius: float,
        start_angle: float,
        end_angle: float,
        layer: str = "0",
        color: str | int = "white",
        lineweight: int = 0,
        _skip_refresh: bool = False,
    ) -> str:
        """Draw an arc.

        Args:
            _skip_refresh: Internal flag to skip view refresh (used for batch operations)
        """
        document = self._get_document("draw_arc")

        center_pt = CADInterface.normalize_coordinate(center)
        center_array = self._to_variant_array(center_pt)

        arc = document.ModelSpace.AddArc(
            center_array,
            radius,
            self._to_radians(start_angle),
            self._to_radians(end_angle),
        )
        self._apply_properties(arc, layer, color, lineweight)
        self._track_entity(arc, "arc")
        if not _skip_refresh:
            self.refresh_view()

        logger.debug(f"Drew arc at {center} from {start_angle}° to {end_angle}°")
        return str(arc.Handle)

    def draw_rectangle(
        self,
        corner1: Coordinate,
        corner2: Coordinate,
        layer: str = "0",
        color: str | int = "white",
        lineweight: int = 0,
        _skip_refresh: bool = False,
    ) -> str:
        """Draw a rectangle from two corners.

        Args:
            _skip_refresh: Internal flag to skip view refresh (used for batch operations)
        """
        self._validate_connection()
        pt1 = CADInterface.normalize_coordinate(corner1)
        pt2 = CADInterface.normalize_coordinate(corner2)

        # Create rectangle corners
        points: List[Coordinate] = [
            (pt1[0], pt1[1], pt1[2]),
            (pt2[0], pt1[1], pt1[2]),
            (pt2[0], pt2[1], pt2[2]),
            (pt1[0], pt2[1], pt2[2]),
            (pt1[0], pt1[1], pt1[2]),  # Close
        ]

        # Use polyline for rectangle
        return self.draw_polyline(
            points,
            closed=True,
            layer=layer,
            color=color,
            lineweight=lineweight,
            _skip_refresh=_skip_refresh,
        )

    def draw_polyline(
        self,
        points: List[Coordinate],
        closed: bool = False,
        layer: str = "0",
        color: str | int = "white",
        lineweight: int = 0,
        _skip_refresh: bool = False,
    ) -> str:
        """Draw a polyline through points.

        Args:
            _skip_refresh: Internal flag to skip view refresh (used for batch operations)
        """
        document = self._get_document("draw_polyline")

        if len(points) < 2:
            raise InvalidParameterError("points", points, "at least 2 points")

        # Convert to 3D points and flatten to variant array
        normalized_points = [CADInterface.normalize_coordinate(p) for p in points]
        variant_points = self._points_to_variant_array(normalized_points)

        polyline = document.ModelSpace.AddPolyline(variant_points)

        if closed:
            polyline.Closed = True

        self._apply_properties(polyline, layer, color, lineweight)
        self._track_entity(polyline, "polyline")
        if not _skip_refresh:
            self.refresh_view()

        logger.debug(f"Drew polyline with {len(points)} points")
        return str(polyline.Handle)

    def draw_ellipse(
        self,
        center: Coordinate,
        major_axis_end: Coordinate,
        minor_axis_ratio: float,
        layer: str = "0",
        color: str | int = "white",
        lineweight: int = 0,
    ) -> str:
        """Draw an ellipse."""
        document = self._get_document("draw_ellipse")

        center_pt = CADInterface.normalize_coordinate(center)
        major_end = CADInterface.normalize_coordinate(major_axis_end)

        center_array = self._to_variant_array(center_pt)
        major_array = self._to_variant_array(major_end)

        ellipse = document.ModelSpace.AddEllipse(
            center_array, major_array, minor_axis_ratio
        )
        self._apply_properties(ellipse, layer, color, lineweight)
        self._track_entity(ellipse, "ellipse")
        self.refresh_view()

        logger.debug(f"Drew ellipse at {center}")
        return str(ellipse.Handle)

    def draw_text(
        self,
        position: Coordinate,
        text: str,
        height: float = 2.5,
        rotation: float = 0.0,
        layer: str = "0",
        color: str | int = "white",
        _skip_refresh: bool = False,
    ) -> str:
        """Add text to drawing.

        Args:
            _skip_refresh: Internal flag to skip view refresh (used for batch operations)
        """
        document = self._get_document("draw_text")

        pos = CADInterface.normalize_coordinate(position)
        pos_array = self._to_variant_array(pos)

        text_obj = document.ModelSpace.AddText(text, pos_array, height)
        text_obj.Rotation = self._to_radians(rotation)

        self._apply_properties(text_obj, layer, color)
        self._track_entity(text_obj, "text")
        if not _skip_refresh:
            self.refresh_view()

        logger.debug(f"Added text '{text}' at {position}")
        return str(text_obj.Handle)

    def draw_hatch(
        self,
        boundary_points: List[Coordinate],
        pattern: str = "SOLID",
        scale: float = 1.0,
        angle: float = 0.0,
        color: str | int = "white",
        layer: str = "0",
    ) -> str:
        """Create a hatch (filled area)."""
        document = self._get_document("draw_hatch")

        # Create boundary polyline (invisible)
        boundary_polyline = document.ModelSpace.AddPolyline(
            self._points_to_variant_array(
                [CADInterface.normalize_coordinate(p) for p in boundary_points]
            )
        )
        boundary_polyline.Closed = True

        # Create hatch
        hatch = document.ModelSpace.AddHatch(
            0, pattern, True
        )  # 0 = Normal, True = Associative
        hatch.AppendOuterLoop([boundary_polyline])
        hatch.Evaluate()

        self._apply_properties(hatch, layer, color)
        self._track_entity(hatch, "hatch")
        self.refresh_view()

        logger.debug(f"Created hatch with pattern {pattern}")
        return str(hatch.Handle)

    def add_dimension(
        self,
        start: Coordinate,
        end: Coordinate,
        text_position: Optional[Coordinate] = None,
        text: Optional[str] = None,
        layer: str = "0",
        color: str | int = "white",
        offset: float = 10.0,
        _skip_refresh: bool = False,
    ) -> str:
        """Add a dimension annotation with optional offset from the edge.

        Args:
            start: Start point of the dimension
            end: End point of the dimension
            text_position: Position for dimension text (optional)
            text: Custom dimension text (optional)
            layer: Layer name
            color: Color name or index
            offset: Distance to offset the dimension line from the edge (default: 10.0)
            _skip_refresh: Internal flag to skip view refresh (used for batch operations)

        Returns:
            Entity handle of the created dimension
        """
        document = self._get_document("add_dimension")

        start_pt = CADInterface.normalize_coordinate(start)
        end_pt = CADInterface.normalize_coordinate(end)

        start_array = self._to_variant_array(start_pt)
        end_array = self._to_variant_array(end_pt)

        # Calculate perpendicular offset point for the dimension line
        dx = end_pt[0] - start_pt[0]
        dy = end_pt[1] - start_pt[1]
        length = math.sqrt(dx * dx + dy * dy)

        if length > 0:
            # Perpendicular to (dx, dy) is (-dy, dx)
            perp_x = -dy / length
            perp_y = dx / length

            # Apply offset in perpendicular direction
            offset_x = perp_x * offset
            offset_y = perp_y * offset

            # Midpoint of the dimension line, offset perpendicularly
            mid_x = (start_pt[0] + end_pt[0]) / 2 + offset_x
            mid_y = (start_pt[1] + end_pt[1]) / 2 + offset_y
            mid_z = start_pt[2]

            dim_position = self._to_variant_array((mid_x, mid_y, mid_z))
        else:
            # If start and end are the same, use default offset
            dim_position = self._to_variant_array(
                (start_pt[0] + offset, start_pt[1], start_pt[2])
            )

        # Use aligned dimension with offset position
        dim = document.ModelSpace.AddDimAligned(start_array, end_array, dim_position)

        if text:
            dim.TextOverride = text

        self._apply_properties(dim, layer, color)
        self._track_entity(dim, "dimension")
        if not _skip_refresh:
            self.refresh_view()

        logger.debug(f"Added dimension from {start} to {end} with offset {offset}")
        return str(dim.Handle)

    # ========== Layer Management ==========

    def create_layer(
        self,
        name: str,
        color: str | int = "white",
        lineweight: int = 0,
    ) -> bool:
        """Create a new layer."""
        try:
            document = self._get_document("create_layer")

            layer_obj = document.Layers.Add(name)

            if isinstance(color, str):
                color = self._get_color_index(color)
            layer_obj.Color = color

            if self.validate_lineweight(lineweight) == lineweight:
                layer_obj.LineWeight = lineweight

            logger.info(f"Created layer '{name}'")
            return True

        except Exception as e:
            logger.error(f"Failed to create layer '{name}': {e}")
            return False

    def set_current_layer(self, name: str) -> bool:
        """Set active layer."""
        try:
            document = self._get_document("set_current_layer")

            document.ActiveLayer = document.Layers.Item(name)
            self._drawing_state["current_layer"] = name
            logger.debug(f"Set current layer to '{name}'")
            return True
        except Exception as e:
            logger.error(f"Failed to set current layer: {e}")
            return False

    def get_current_layer(self) -> str:
        """Get current active layer."""
        try:
            document = self._get_document("get_current_layer")
            return str(document.ActiveLayer.Name)
        except Exception:
            current_layer = self._drawing_state["current_layer"]
            return str(current_layer) if current_layer else "0"

    def list_layers(self) -> List[str]:
        """Get list of all layers."""
        try:
            document = self._get_document("list_layers")
            layers: List[str] = []
            for layer in document.Layers:
                layers.append(layer.Name)
            return layers
        except Exception as e:
            logger.error(f"Failed to list layers: {e}")
            return []

    def get_layers_info(
        self, entity_data: Optional[List[Dict[str, Any]]] = None
    ) -> List[Dict[str, Any]]:
        """Get detailed information about all layers.

        Optimized to count entities per layer in a single pass using direct iteration,
        or from pre-extracted entity data to avoid re-iterating ModelSpace.

        Args:
            entity_data: Optional pre-extracted entity data. If provided, layer counts
                        will be computed from this data instead of iterating ModelSpace.

        Returns:
            List of dictionaries with layer information:
            - Name: Layer name
            - ObjectCount: Number of objects on the layer
            - Color: Layer color
            - Linetype: Layer linetype
            - Lineweight: Layer lineweight
            - IsLocked: Whether layer is locked
            - IsVisible: Whether layer is visible
        """
        try:
            document = self._get_document("get_layers_info")
            layers_info = []

            # OPTIMIZATION: Use pre-extracted data if available to avoid re-iteration
            layer_counts: Dict[str, int] = {}

            if entity_data is not None:
                # Count from pre-extracted data (MUCH faster - no COM calls)
                logger.debug(
                    f"Computing layer counts from {len(entity_data)} pre-extracted entities"
                )
                for entity in entity_data:
                    layer_name = entity.get("Layer", "0")
                    layer_counts[layer_name] = layer_counts.get(layer_name, 0) + 1
            else:
                # Fallback: iterate ModelSpace (slower - requires COM calls)
                logger.debug("Iterating ModelSpace to count entities by layer")
                model_space = document.ModelSpace

                try:
                    # Direct iteration (faster than Item(i))
                    for entity in model_space:
                        try:
                            layer_name = self._fast_get_property(entity, "Layer", "0")
                            layer_counts[layer_name] = (
                                layer_counts.get(layer_name, 0) + 1
                            )
                        except Exception:
                            pass
                except Exception:
                    # Fallback to indexed iteration if direct iteration fails
                    try:
                        entity_count = model_space.Count
                        for i in range(entity_count):
                            try:
                                entity = model_space.Item(i)
                                layer_name = self._fast_get_property(
                                    entity, "Layer", "0"
                                )
                                layer_counts[layer_name] = (
                                    layer_counts.get(layer_name, 0) + 1
                                )
                            except Exception:
                                pass
                    except Exception as e:
                        logger.debug(f"Failed to count entities by layer: {e}")

            # Build layer information
            for layer in document.Layers:
                try:
                    # Get layer properties
                    layer_color = self._safe_get_property(layer, "Color", 256)
                    color_map_reverse = {v: k for k, v in COLOR_MAP.items()}
                    color_name = color_map_reverse.get(layer_color, str(layer_color))

                    layer_info = {
                        "Name": str(layer.Name),
                        "ObjectCount": layer_counts.get(str(layer.Name), 0),
                        "Color": color_name,
                        "Linetype": str(
                            self._safe_get_property(layer, "Linetype", "Continuous")
                        ),
                        "Lineweight": str(
                            self._safe_get_property(layer, "Lineweight", "Default")
                        ),
                        "IsLocked": bool(self._safe_get_property(layer, "Lock", False)),
                        "IsVisible": not bool(
                            self._safe_get_property(layer, "Frozen", False)
                        ),
                    }
                    layers_info.append(layer_info)
                except Exception as e:
                    logger.debug(f"Failed to get info for layer {layer.Name}: {e}")
                    continue

            return layers_info
        except Exception as e:
            logger.error(f"Failed to get layers info: {e}")
            return []

    def rename_layer(self, old_name: str, new_name: str) -> bool:
        """Rename an existing layer."""
        try:
            self._validate_connection()
            document = self._get_document("rename_layer")

            if old_name == "0":
                logger.error("Cannot rename layer '0' (standard layer)")
                return False

            layer = document.Layers.Item(old_name)
            layer.Name = new_name
            logger.info(f"Renamed layer '{old_name}' to '{new_name}'")
            return True
        except Exception as e:
            logger.error(f"Failed to rename layer '{old_name}' to '{new_name}': {e}")
            return False

    def delete_layer(self, name: str) -> bool:
        """Delete a layer from the drawing."""
        try:
            self._validate_connection()
            document = self._get_document("delete_layer")

            if name == "0":
                logger.error("Cannot delete layer '0' (standard layer)")
                return False

            layer = document.Layers.Item(name)
            layer.Delete()
            logger.info(f"Deleted layer '{name}'")
            return True
        except Exception as e:
            logger.error(f"Failed to delete layer '{name}': {e}")
            return False

    def turn_layer_on(self, name: str) -> bool:
        """Turn on (make visible) a layer."""
        try:
            self._validate_connection()
            document = self._get_document("turn_layer_on")

            layer = document.Layers.Item(name)
            layer.Freeze = False
            logger.info(f"Turned on layer '{name}'")
            return True
        except Exception as e:
            logger.error(f"Failed to turn on layer '{name}': {e}")
            return False

    def turn_layer_off(self, name: str) -> bool:
        """Turn off (hide) a layer."""
        try:
            self._validate_connection()
            document = self._get_document("turn_layer_off")

            layer = document.Layers.Item(name)
            layer.Freeze = True
            logger.info(f"Turned off layer '{name}'")
            return True
        except Exception as e:
            logger.error(f"Failed to turn off layer '{name}': {e}")
            return False

    def is_layer_on(self, name: str) -> bool:
        """Check if a layer is visible (turned on)."""
        try:
            self._validate_connection()
            document = self._get_document("is_layer_on")

            layer = document.Layers.Item(name)
            return not layer.Freeze
        except Exception as e:
            logger.error(f"Failed to check layer '{name}' visibility: {e}")
            return False

    def set_layer_color(self, layer_name: str, color: str | int) -> bool:
        """Set the color of a layer.

        Args:
            layer_name: Name of the layer to modify
            color: Color name (from COLOR_MAP) or ACI index (1-255)

        Returns:
            bool: True if successful, False otherwise

        Note:
            - Uses the modern TrueColor property (recommended by Autodesk)
            - Accepts color names: "red", "blue", "green", etc.
            - Accepts ACI index: 1-255
            - Color value 256 (bylayer) is not valid for layers
        """
        try:
            self._validate_connection()
            document = self._get_document("set_layer_color")
            app = self._get_application("set_layer_color")

            # Get the layer
            try:
                layer = document.Layers.Item(layer_name)
            except Exception:
                raise LayerError(layer_name, "Layer does not exist")

            # Convert color name to ACI index
            if isinstance(color, str):
                color_index = self._get_color_index(color)
            else:
                color_index = color

            # Validate color index (1-255 for layers, 256 is not valid)
            if color_index == 256:
                raise ColorError(
                    str(color_index),
                    "Color 'bylayer' (256) is not valid for layers. Use a specific ACI color (1-255).",
                )
            if not (0 <= color_index <= 255):
                raise ColorError(
                    str(color_index), "Invalid color index. Must be 0-255."
                )

            # Create AcCmColor object (modern method)
            color_obj = app.GetInterfaceObject("AutoCAD.AcCmColor.20")
            color_obj.ColorIndex = color_index

            # Apply to layer using TrueColor property
            layer.TrueColor = color_obj

            logger.info(f"Set layer '{layer_name}' color to ACI {color_index}")
            return True

        except (LayerError, ColorError):
            raise
        except Exception as e:
            logger.error(f"Failed to set layer '{layer_name}' color: {e}")
            return False

    def set_entities_color_bylayer(self, handles: List[str]) -> Dict[str, Any]:
        """Set entities to use their layer's color (ByLayer).

        Args:
            handles: List of entity handles to modify

        Returns:
            dict: Result summary with counts and details:
                - total: Total entities processed
                - changed: Number successfully changed to ByLayer
                - failed: Number that failed
                - results: List of per-entity results

        Note:
            - Assigns color value 256 (acByLayer) to entities
            - Entities will inherit their layer's color
            - Uses TrueColor property (modern method)
        """
        try:
            self._validate_connection()
            document = self._get_document("set_entities_color_bylayer")
            app = self._get_application("set_entities_color_bylayer")
            model_space = document.ModelSpace

            results = []
            changed_count = 0
            failed_count = 0

            # Create AcCmColor object for ByLayer (256)
            color_obj = app.GetInterfaceObject("AutoCAD.AcCmColor.20")
            color_obj.ColorIndex = 256  # acByLayer

            for handle in handles:
                try:
                    # Get entity by handle
                    try:
                        entity = model_space.Item(model_space.Count - 1)
                        for i in range(model_space.Count):
                            entity = model_space.Item(i)
                            if str(entity.Handle) == str(handle):
                                break
                        else:
                            results.append(
                                {
                                    "handle": handle,
                                    "success": False,
                                    "error": "Entity not found",
                                }
                            )
                            failed_count += 1
                            continue
                    except Exception as e:
                        results.append(
                            {
                                "handle": handle,
                                "success": False,
                                "error": f"Failed to find entity: {e}",
                            }
                        )
                        failed_count += 1
                        continue

                    # Set color to ByLayer using TrueColor
                    entity.TrueColor = color_obj

                    results.append({"handle": handle, "success": True})
                    changed_count += 1

                except Exception as e:
                    results.append(
                        {"handle": handle, "success": False, "error": str(e)}
                    )
                    failed_count += 1

            logger.info(f"Set {changed_count}/{len(handles)} entities to ByLayer color")

            return {
                "total": len(handles),
                "changed": changed_count,
                "failed": failed_count,
                "results": results,
            }

        except Exception as e:
            logger.error(f"Failed to set entities to ByLayer: {e}")
            return {
                "total": len(handles),
                "changed": 0,
                "failed": len(handles),
                "error": str(e),
                "results": [],
            }

    # ========== File Operations ==========

    def save_drawing(
        self, filepath: str = "", filename: str = "", format: str = "dwg"
    ) -> bool:
        """Save drawing to file.

        Args:
            filepath: Full path to save file (e.g., 'C:/drawings/myfile.dwg')
            filename: Just the filename (e.g., 'myfile.dwg'). If provided without
                     filepath, uses configured output directory
            format: File format (dwg, dxf, etc.). Default: dwg

        Returns:
            bool: True if successful, False otherwise

        Note:
            - If both filepath and filename provided, filepath takes precedence
            - If only filename provided, saved to config output directory
            - If neither provided, uses current document name
        """
        try:
            document = self._get_document("save_drawing")
            config = get_config()

            # SECURITY: Resolve output directory first (reference for validation)
            output_dir = Path(config.output.directory).expanduser().resolve()

            # ========== Determine Directory (filepath) ==========
            if filepath:
                # If filepath provided, extract directory part
                dir_part = str(Path(filepath).parent)
                if dir_part and dir_part != ".":
                    save_dir = dir_part
                else:
                    save_dir = str(output_dir)
            else:
                save_dir = str(output_dir)

            # Convert to absolute path (required by AutoCAD COM API)
            save_dir_path = Path(save_dir).expanduser().resolve()

            # SECURITY: Verify the directory is within the configured output directory
            try:
                save_dir_path.relative_to(output_dir)
            except ValueError:
                logger.error(
                    f"Security: Attempted to save outside output directory. "
                    f"Requested: {save_dir_path}, Allowed: {output_dir}"
                )
                raise CADOperationError(
                    "save_drawing",
                    f"File path must be within {output_dir}",
                )

            # Create directory if it doesn't exist
            save_dir_path.mkdir(parents=True, exist_ok=True)

            # ========== Determine Filename ==========
            if filepath:
                file_part = str(Path(filepath).name)
                if file_part and file_part != ".":
                    save_filename = file_part
                elif filename:
                    save_filename = filename
                else:
                    save_filename = None
            else:
                save_filename = filename

            # If still no filename, use document name or generate one
            if not save_filename:
                if document.Name:
                    save_filename = document.Name
                else:
                    from datetime import datetime

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    save_filename = f"drawing_{timestamp}.{format}"

            # ========== Ensure Correct File Extension ==========
            if not save_filename.lower().endswith(f".{format}"):
                save_filename = f"{save_filename}.{format}"

            # ========== Combine Directory + Filename ==========
            final_path = save_dir_path / save_filename

            # SECURITY: Final validation - ensure combined path is still within output directory
            try:
                final_path.resolve().relative_to(output_dir)
            except ValueError:
                logger.error(
                    f"Security: Final path validation failed. "
                    f"Path: {final_path}, Allowed dir: {output_dir}"
                )
                raise CADOperationError(
                    "save_drawing",
                    f"Invalid file path: {final_path}",
                )

            # Save the drawing
            document.SaveAs(str(final_path))
            logger.info(f"Saved drawing to {final_path}")
            return True
        except Exception as e:
            logger.error(f"Failed to save drawing: {e}")
            return False

    def open_drawing(self, filepath: str) -> bool:
        """Open a drawing file."""
        try:
            application = self._get_application("open_drawing")
            self.document = application.Documents.Open(filepath)
            logger.info(f"Opened drawing from {filepath}")
            return True
        except Exception as e:
            logger.error(f"Failed to open drawing: {e}")
            return False

    def new_drawing(self) -> bool:
        """Create new blank drawing."""
        try:
            application = self._get_application("new_drawing")
            self.document = application.Documents.Add()
            self._refresh_document_reference()
            logger.info("Created new blank drawing")
            return True
        except Exception as e:
            logger.error(f"Failed to create new drawing: {e}")
            return False

    def _refresh_document_reference(self, auto_create: bool = True) -> bool:
        """Refresh internal document reference to ActiveDocument.

        This ensures self.document always points to the active document
        in the application. Useful after creating or switching documents.

        Args:
            auto_create: If True and no documents open, create a new one (default: True)

        Returns:
            True if successful, False otherwise
        """
        try:
            application = self._get_application("_refresh_document_reference")

            # Case 1: Documents are open, use the active one
            if application.Documents.Count > 0:
                self.document = application.ActiveDocument
                if self.document is not None:
                    logger.debug(f"Document reference refreshed: {self.document.Name}")
                return True

            # Case 2: No documents open
            if auto_create:
                logger.warning("No documents open. Creating a new blank document...")
                self.document = application.Documents.Add()
                if self.document is not None:
                    logger.info(f"Auto-created new document: {self.document.Name}")
                return True
            else:
                logger.warning("No documents open")
                return False

        except Exception as e:
            logger.error(f"Failed to refresh document reference: {e}")
            return False

    def get_open_drawings(self) -> list:
        """Get list of all open drawing filenames.

        Returns:
            List of drawing names (e.g., ["drawing1.dwg", "drawing2.dwg"])
        """
        try:
            application = self._get_application("get_open_drawings")
            drawings = []

            # Use direct iteration instead of Item indexing
            for doc in application.Documents:
                drawings.append(doc.Name)

            logger.info(f"Found {len(drawings)} open drawings: {drawings}")
            return drawings
        except Exception as e:
            logger.error(f"Failed to get open drawings: {e}")
            return []

    def switch_drawing(self, drawing_name: str) -> bool:
        """Switch to a different open drawing.

        Args:
            drawing_name: Name of the drawing to switch to (e.g., "drawing1.dwg")

        Returns:
            True if successful, False otherwise
        """
        try:
            application = self._get_application("switch_drawing")

            # Use direct iteration instead of Item indexing
            for doc in application.Documents:
                if doc.Name == drawing_name:
                    doc.Activate()
                    self.document = doc
                    logger.info(f"Switched to drawing: {drawing_name}")
                    return True

            logger.warning(f"Drawing not found: '{drawing_name}'")
            return False

        except Exception as e:
            logger.error(f"Failed to switch drawing: {e}")
            return False

    def close_drawing(self, save_changes: bool = False) -> bool:
        """Close the current drawing.

        Args:
            save_changes: Whether to save changes before closing (default: False)
                         True = save changes
                         False = discard changes without prompting

        Returns:
            True if successful, False otherwise
        """
        try:
            if not self._validate_document() or self.document is None:
                logger.warning("No document to close")
                return False

            document = self.document
            doc_name = document.Name

            # Close document using COM API
            document.Close(save_changes)

            # Try to update connection to remaining open document
            refresh_success = self._refresh_document_reference(auto_create=False)

            if refresh_success and self.document is not None:
                logger.info(
                    f"Closed drawing: {doc_name} (save_changes={save_changes}). "
                    f"Switched to: {self.document.Name}"
                )
            else:
                # No other documents open - attempt to create one to maintain connection
                logger.warning(
                    f"No other documents open after closing {doc_name}. "
                    "Attempting to create a new blank document..."
                )
                try:
                    application = self._get_application("close_drawing_reconnect")
                    self.document = application.Documents.Add()
                    logger.info(
                        f"Closed drawing: {doc_name} (save_changes={save_changes}). "
                        f"Created new document: {self.document.Name}"
                    )
                except Exception as e:
                    self.document = None
                    logger.info(
                        f"Closed drawing: {doc_name} (save_changes={save_changes}). "
                        "Could not create new document."
                    )
                    logger.debug(f"Auto-create error: {e}")

            return True

        except Exception as e:
            logger.error(f"Failed to close drawing: {e}")
            return False

    # ========== View Management ==========

    def zoom_extents(self) -> bool:
        """Zoom to show all entities."""
        try:
            application = self._get_application("zoom_extents")
            application.ZoomExtents()
            logger.debug("Zoomed to extents")
            return True
        except Exception as e:
            logger.error(f"Failed to zoom extents: {e}")
            return False

    def refresh_view(self) -> bool:
        """Refresh the view using multiple techniques for maximum compatibility.

        Uses a combination of techniques in fallback order:
        1. Application.Refresh() (COM API - no undo/redo impact)
        2. SendCommand with REDRAW (most reliable visual update)
        3. Window click simulation (forces UI update)

        Note: REDRAW command is not wrapped in UNDO to avoid complicating
        the undo/redo stack. If refresh_view is called during user operations,
        the REDRAW will be undone by the user's undo command anyway.

        Returns:
            True if refresh was attempted (best effort approach)
        """
        try:
            application = self._get_application("refresh_view")
            document = self._get_document("refresh_view")

            # Technique 1: COM API Refresh (doesn't affect undo/redo)
            try:
                application.Refresh()
                logger.debug("Refresh: COM Refresh executed")
            except Exception as e:
                logger.debug(f"COM Refresh failed: {e}")

            # Technique 2: Send REDRAW command (most reliable visual update)
            try:
                document.SendCommand("_redraw\n")
                logger.debug("Refresh: REDRAW command sent")
            except Exception as e:
                logger.debug(f"REDRAW command failed: {e}")

            # Technique 3: Simulate click on CAD window (forces UI update)
            self._simulate_autocad_click()

            return True
        except Exception as e:
            logger.debug(f"refresh_view error: {e}")
            return False

    def _simulate_autocad_click(self) -> bool:
        """Simulate a click in the CAD window to force viewport update.

        This is a workaround to ensure the viewport updates after operations.
        Finds the CAD main window and simulates a subtle click.

        Returns:
            True if click simulation succeeded, False otherwise
        """
        try:
            self._validate_connection()

            hwnd = None
            for class_name in AUTOCAD_WINDOW_CLASSES:
                hwnd = win32gui.FindWindow(class_name, None)
                if hwnd:
                    logger.debug(f"Found CAD window: {class_name}")
                    break

            if not hwnd:
                logger.debug("CAD window not found for click simulation")
                return False

            # Get window center position for subtle click
            try:
                rect = win32gui.GetWindowRect(hwnd)
                x = (rect[0] + rect[2]) // 2  # Center X
                y = (rect[1] + rect[3]) // 2  # Center Y

                # Simulate left mouse click at window center
                win32api.SetCursorPos((x, y))
                time.sleep(CLICK_DELAY / 1000.0)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x, y, 0, 0)
                time.sleep(CLICK_HOLD_DELAY / 1000.0)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x, y, 0, 0)

                logger.debug("CAD window click simulated")
                return True
            except Exception as e:
                logger.debug(f"Click simulation failed: {e}")
                return False

        except Exception as e:
            logger.debug(f"_simulate_autocad_click error: {e}")
            return False

    # ========== Entity Management ==========

    def delete_entity(self, handle: str) -> bool:
        """Delete entity by handle."""
        try:
            document = self._get_document("delete_entity")

            entity = document.HandleToObject(handle)
            entity.Delete()
            logger.debug(f"Deleted entity {handle}")
            return True
        except Exception as e:
            logger.error(f"Failed to delete entity: {e}")
            return False

    def get_entity_properties(self, handle: str) -> Dict[str, Any]:
        """Get entity properties."""
        try:
            document = self._get_document("get_entity_properties")

            entity = document.HandleToObject(handle)
            return {
                "handle": entity.Handle,
                "object_name": entity.ObjectName,
                "layer": entity.Layer,
                "color": entity.Color,
                "lineweight": entity.LineWeight,
            }
        except Exception as e:
            logger.error(f"Failed to get entity properties: {e}")
            return {}

    def set_entity_properties(self, handle: str, properties: Dict[str, Any]) -> bool:
        """Modify entity properties."""
        try:
            document = self._get_document("set_entity_properties")

            entity = document.HandleToObject(handle)

            if "layer" in properties:
                entity.Layer = properties["layer"]
            if "color" in properties:
                color = properties["color"]
                if isinstance(color, str):
                    color = self._get_color_index(color)
                entity.Color = color
            if "lineweight" in properties:
                entity.LineWeight = self.validate_lineweight(properties["lineweight"])

            logger.debug(f"Updated properties for entity {handle}")
            return True
        except Exception as e:
            logger.error(f"Failed to set entity properties: {e}")
            return False

    # ========== Entity Selection ==========

    def _select_entities_generic(
        self,
        filter_func: Callable[[Any], bool],
        selection_set_name: str,
        description: str,
    ) -> List[str]:
        """Generic entity selection helper.

        Args:
            filter_func: Function that takes an entity and returns True if it matches criteria
            selection_set_name: Name for the selection set
            description: Description for logging

        Returns:
            List of entity handles that match criteria
        """
        try:
            self._validate_connection()
            document = self._get_document("select")
            app = self._get_application("select")

            selected_handles = []
            entities_to_select = []

            # Iterate through all entities in ModelSpace
            for entity in document.ModelSpace:
                try:
                    if filter_func(entity):
                        handle = str(entity.Handle)
                        selected_handles.append(handle)
                        entities_to_select.append(entity)
                        logger.debug(f"Found {description}: {handle}")
                except Exception as e:
                    logger.debug(f"Error processing entity: {e}")
                    continue

            # Create visible selection using SelectionSet
            if entities_to_select:
                try:
                    # Clear current selection
                    app.ActiveDocument.Select(-1)
                    self._delete_selection_set(document, selection_set_name)

                    ss = document.SelectionSets.Add(selection_set_name)
                    for entity in entities_to_select:
                        try:
                            ss.Select(SELECTION_SET_IMPLIED, None, entity)
                            logger.debug(
                                f"Added entity {entity.Handle} to selection set"
                            )
                        except Exception as e:
                            logger.warning(
                                f"Failed to add entity to selection set: {e}"
                            )
                except Exception as e:
                    logger.warning(f"Failed to create selection set: {e}")

            logger.info(f"Selected {len(selected_handles)} {description}")
            return selected_handles

        except Exception as e:
            logger.error(f"Failed to select {description}: {e}")
            return []

    def select_by_color(self, color: str | int) -> List[str]:
        """Select all entities of a specific color."""
        if isinstance(color, str):
            color = self._get_color_index(color)

        def color_filter(entity: Any) -> bool:
            return hasattr(entity, "Color") and entity.Color == color

        return self._select_entities_generic(
            color_filter, SS_COLOR_SELECT, f"entities with color {color}"
        )

    def select_by_layer(self, layer_name: str) -> List[str]:
        """Select all entities on a specific layer."""
        target_layer = layer_name.strip()

        def layer_filter(entity: Any) -> bool:
            try:
                # Get layer name - try multiple approaches
                entity_layer = None
                try:
                    entity_layer = str(entity.Layer).strip()
                except Exception:
                    try:
                        entity_layer = str(
                            entity.Properties.Item("Layer").Value
                        ).strip()
                    except Exception:
                        return False

                # Normalize and case-insensitive comparison
                return entity_layer.lower() == target_layer.lower()
            except Exception:
                return False

        return self._select_entities_generic(
            layer_filter, SS_LAYER_SELECT, f"entities on layer '{layer_name}'"
        )

    def select_by_type(self, entity_type: str) -> List[str]:
        """Select all entities of a specific type."""
        # Map user-friendly types to AutoCAD object names
        type_map = {
            "line": "AcDbLine",
            "circle": "AcDbCircle",
            "arc": "AcDbArc",
            "polyline": "AcDb2dPolyline",
            "text": "AcDbText",
            "point": "AcDbPoint",
        }

        object_name = type_map.get(entity_type.lower(), entity_type)
        logger.debug(f"Searching for entities of type: {object_name}")

        def type_filter(entity: Any) -> bool:
            try:
                current_object_name = entity.ObjectName
                return (
                    current_object_name == object_name
                    or entity_type.lower() in current_object_name.lower()
                )
            except Exception:
                return False

        return self._select_entities_generic(
            type_filter, SS_TYPE_SELECT, f"entities of type '{entity_type}'"
        )

    def get_selected_entities(self) -> List[str]:
        """Get list of currently selected entities."""
        try:
            self._validate_connection()
            app = self._get_application("get_selected_entities")
            selected = app.ActiveDocument.SelectionSets.Add(SS_SELECTION_GET)

            handles = []
            try:
                for entity in selected:
                    handles.append(str(entity.Handle))
            finally:
                selected.Delete()

            logger.debug(f"Got {len(handles)} selected entities")
            return handles
        except Exception as e:
            logger.error(f"Failed to get selected entities: {e}")
            return []

    def clear_selection(self) -> bool:
        """Clear current selection."""
        try:
            self._validate_connection()
            app = self._get_application("clear_selection")
            app.ActiveDocument.Select(-1)  # Select nothing
            logger.debug("Selection cleared")
            return True
        except Exception as e:
            logger.error(f"Failed to clear selection: {e}")
            return False

    # ========== Block Creation ==========

    def create_block_from_entities(
        self,
        block_name: str,
        entity_handles: List[str],
        insertion_point: Coordinate = (0.0, 0.0, 0.0),
        description: str = "",
    ) -> Dict[str, Any]:
        """Create a block from specified entities.

        Args:
            block_name: Name for the new block
            entity_handles: List of entity handles to include in block
            insertion_point: Base point for block definition (default: 0,0,0)
            description: Optional block description

        Returns:
            Dictionary with operation status and details

        Raises:
            CADOperationError: If block creation fails
            InvalidParameterError: If parameters are invalid
        """
        try:
            self._validate_connection()
            app = self._get_application("create_block_from_entities")
            document = app.ActiveDocument

            # Validate block name
            if not block_name or not isinstance(block_name, str):
                raise InvalidParameterError(
                    "block_name", block_name, "Block name must be a non-empty string"
                )

            # Check if block already exists
            try:
                existing_block = document.Blocks.Item(block_name)
                # If we get here, block exists
                raise CADOperationError(
                    "create_block",
                    f"Block '{block_name}' already exists. Choose a different name.",
                )
            except CADOperationError:
                # Re-raise our error
                raise
            except Exception:
                # Block doesn't exist (Item() raised exception), continue
                pass

            # Convert insertion point
            insert_pt = CADInterface.normalize_coordinate(insertion_point)
            insert_pt_variant = self._to_variant_array(insert_pt)

            # Create block definition
            block_def = document.Blocks.Add(insert_pt_variant, block_name)

            # Set description if provided
            if description:
                try:
                    block_def.Comments = description
                except Exception as e:
                    logger.warning(f"Could not set block description: {e}")

            # Get entities from handles
            entities = []
            failed_handles = []
            for handle in entity_handles:
                try:
                    entity = document.HandleToObject(handle)
                    entities.append(entity)
                except Exception as e:
                    logger.warning(f"Could not get entity with handle {handle}: {e}")
                    failed_handles.append(handle)

            if not entities:
                raise CADOperationError(
                    "create_block",
                    f"No valid entities found from {len(entity_handles)} handles provided",
                )

            # Convert entities to variant array
            entities_variant = self._objects_to_variant_array(entities)

            # Copy entities to block definition
            try:
                document.CopyObjects(entities_variant, block_def)
            except Exception as e:
                raise CADOperationError(
                    "create_block",
                    f"Failed to copy entities to block: {str(e)}",
                )

            logger.info(
                f"Created block '{block_name}' with {len(entities)} entities "
                f"at insertion point {insert_pt}"
            )

            return {
                "success": True,
                "block_name": block_name,
                "total_handles": len(entity_handles),
                "entities_added": len(entities),
                "failed_handles": failed_handles,
                "insertion_point": insert_pt,
            }

        except (CADOperationError, InvalidParameterError):
            raise
        except Exception as e:
            logger.error(f"Failed to create block from entities: {e}")
            raise CADOperationError("create_block", str(e))

    def create_block_from_selection(
        self,
        block_name: str,
        insertion_point: Coordinate = (0.0, 0.0, 0.0),
        description: str = "",
    ) -> Dict[str, Any]:
        """Create a block from currently selected entities.

        Args:
            block_name: Name for the new block
            insertion_point: Base point for block definition (default: 0,0,0)
            description: Optional block description

        Returns:
            Dictionary with operation status and details

        Raises:
            CADOperationError: If block creation fails or no entities selected
            InvalidParameterError: If parameters are invalid
        """
        try:
            self._validate_connection()
            app = self._get_application("create_block_from_selection")
            document = app.ActiveDocument

            # Validate block name
            if not block_name or not isinstance(block_name, str):
                raise InvalidParameterError(
                    "block_name", block_name, "Block name must be a non-empty string"
                )

            # Check if block already exists
            try:
                existing_block = document.Blocks.Item(block_name)
                # If we get here, block exists
                raise CADOperationError(
                    "create_block",
                    f"Block '{block_name}' already exists. Choose a different name.",
                )
            except CADOperationError:
                # Re-raise our error
                raise
            except Exception:
                # Block doesn't exist (Item() raised exception), continue
                pass

            # Get currently selected entities
            try:
                selection_set = document.PickfirstSelectionSet
                entity_count = selection_set.Count

                if entity_count == 0:
                    raise CADOperationError(
                        "create_block",
                        "No entities selected. Please select entities in the drawing first.",
                    )

                # Convert selection to list of entities
                entities = []
                for i in range(entity_count):
                    entities.append(selection_set.Item(i))

                logger.debug(f"Retrieved {len(entities)} entities from selection")

            except Exception as e:
                raise CADOperationError(
                    "create_block",
                    f"Failed to get selected entities: {str(e)}",
                )

            # Convert insertion point
            insert_pt = CADInterface.normalize_coordinate(insertion_point)
            insert_pt_variant = self._to_variant_array(insert_pt)

            # Create block definition
            block_def = document.Blocks.Add(insert_pt_variant, block_name)

            # Set description if provided
            if description:
                try:
                    block_def.Comments = description
                except Exception as e:
                    logger.warning(f"Could not set block description: {e}")

            # Convert entities to variant array
            entities_variant = self._objects_to_variant_array(entities)

            # Copy entities to block definition
            try:
                document.CopyObjects(entities_variant, block_def)
            except Exception as e:
                raise CADOperationError(
                    "create_block",
                    f"Failed to copy entities to block: {str(e)}",
                )

            logger.info(
                f"Created block '{block_name}' from {len(entities)} selected entities "
                f"at insertion point {insert_pt}"
            )

            return {
                "success": True,
                "block_name": block_name,
                "entities_added": len(entities),
                "insertion_point": insert_pt,
            }

        except (CADOperationError, InvalidParameterError):
            raise
        except Exception as e:
            logger.error(f"Failed to create block from selection: {e}")
            raise CADOperationError("create_block", str(e))

    # ========== Entity Manipulation ==========

    def move_entities(
        self, handles: List[str], offset_x: float, offset_y: float
    ) -> bool:
        """Move entities by an offset."""
        try:
            self._validate_connection()
            document = self._get_document("move_entities")

            moved_count = 0
            for handle in handles:
                try:
                    entity = document.HandleToObject(handle)

                    from_point = self._to_variant_array((0.0, 0.0, 0.0))
                    to_point = self._to_variant_array((offset_x, offset_y, 0.0))

                    entity.Move(from_point, to_point)
                    moved_count += 1
                    logger.debug(f"Moved entity {handle} by ({offset_x}, {offset_y})")

                except Exception as e:
                    logger.warning(f"Failed to move entity {handle}: {e}")

            logger.info(f"Moved {moved_count}/{len(handles)} entities")
            self.refresh_view()
            return moved_count > 0
        except Exception as e:
            logger.error(f"Failed to move entities: {e}")
            return False

    def rotate_entities(
        self, handles: List[str], center_x: float, center_y: float, angle: float
    ) -> bool:
        """Rotate entities around a point."""
        try:
            self._validate_connection()
            document = self._get_document("rotate_entities")

            rotated_count = 0
            for handle in handles:
                try:
                    entity = document.HandleToObject(handle)
                    center_point = self._to_variant_array((center_x, center_y, 0.0))
                    radians = self._to_radians(angle)

                    entity.Rotate(center_point, radians)
                    rotated_count += 1
                    logger.debug(f"Rotated entity {handle} by {angle}°")

                except Exception as e:
                    logger.warning(f"Failed to rotate entity {handle}: {e}")

            logger.info(f"Rotated {rotated_count}/{len(handles)} entities")
            self.refresh_view()
            return rotated_count > 0
        except Exception as e:
            logger.error(f"Failed to rotate entities: {e}")
            return False

    def scale_entities(
        self, handles: List[str], center_x: float, center_y: float, scale_factor: float
    ) -> bool:
        """Scale entities around a point."""
        try:
            self._validate_connection()
            document = self._get_document("scale_entities")

            scaled_count = 0
            for handle in handles:
                try:
                    entity = document.HandleToObject(handle)
                    center_point = self._to_variant_array((center_x, center_y, 0.0))
                    entity.ScaleEntity(center_point, scale_factor)
                    scaled_count += 1
                    logger.debug(f"Scaled entity {handle} by {scale_factor}")

                except Exception as e:
                    logger.warning(f"Failed to scale entity {handle}: {e}")

            logger.info(f"Scaled {scaled_count}/{len(handles)} entities")
            self.refresh_view()
            return scaled_count > 0
        except Exception as e:
            logger.error(f"Failed to scale entities: {e}")
            return False

    def copy_entities(self, handles: List[str]) -> bool:
        """Copy entities to clipboard using SendCommand."""
        try:
            self._validate_connection()
            document = self._get_document("copy_entities")
            app = self._get_application("copy_entities")

            # Create a selection set with entities to copy
            try:
                self._delete_selection_set(document, SS_COPY)
            except Exception:
                pass

            ss = document.SelectionSets.Add(SS_COPY)
            try:
                for handle in handles:
                    entity = document.HandleToObject(handle)
                    ss.Select(SELECTION_SET_IMPLIED, None, entity)

                # Use SendCommand to execute COPY command
                app.ActiveDocument.SendCommand("_copy\n")
                time.sleep(CLIPBOARD_DELAY / 1000.0)

                logger.info(f"Copied {len(handles)} entities to clipboard")
                return True
            finally:
                self._delete_selection_set(document, SS_COPY)
        except Exception as e:
            logger.error(f"Failed to copy entities: {e}")
            return False

    def paste_entities(self, base_point_x: float, base_point_y: float) -> List[str]:
        """Paste entities from clipboard."""
        try:
            self._validate_connection()
            document = self._get_document("paste_entities")
            app = self._get_application("paste_entities")

            # Get count before paste
            count_before = sum(1 for _ in document.ModelSpace)

            # Paste using SendCommand (more reliable)
            app.ActiveDocument.SendCommand("^V\n")
            time.sleep(CLIPBOARD_STABILITY_DELAY / 1000.0)

            # Get new entities (simplified approach)
            count_after = sum(1 for _ in document.ModelSpace)
            logger.info(f"Pasted {count_after - count_before} entities")

            return []  # Return empty list as we can't reliably track new entities
        except Exception as e:
            logger.error(f"Failed to paste entities: {e}")
            return []

    def change_entity_color(self, handles: List[str], color: str | int) -> bool:
        """Change color of entities."""
        try:
            self._validate_connection()
            document = self._get_document("change_entity_color")

            if isinstance(color, str):
                color = self._get_color_index(color)

            changed_count = 0
            for handle in handles:
                try:
                    entity = document.HandleToObject(handle)
                    entity.Color = color
                    changed_count += 1
                except Exception as e:
                    logger.warning(f"Failed to change color of entity {handle}: {e}")

            logger.info(f"Changed color of {changed_count}/{len(handles)} entities")
            self.refresh_view()
            return changed_count > 0
        except Exception as e:
            logger.error(f"Failed to change entity color: {e}")
            return False

    def change_entity_layer(self, handles: List[str], layer_name: str) -> bool:
        """Move entities to a different layer."""
        try:
            self._validate_connection()
            document = self._get_document("change_entity_layer")

            # Ensure layer exists
            try:
                document.Layers.Item(layer_name)
            except Exception:
                logger.warning(f"Layer '{layer_name}' not found, creating it")
                document.Layers.Add(layer_name)

            changed_count = 0
            for handle in handles:
                try:
                    entity = document.HandleToObject(handle)
                    entity.Layer = layer_name
                    changed_count += 1
                except Exception as e:
                    logger.warning(f"Failed to change layer of entity {handle}: {e}")

            logger.info(
                f"Moved {changed_count}/{len(handles)} entities to layer '{layer_name}'"
            )
            self.refresh_view()
            return changed_count > 0
        except Exception as e:
            logger.error(f"Failed to change entity layer: {e}")
            return False

    # ========== Selection Detection ==========

    def has_selection(self) -> bool:
        """Check if any entities are currently selected.

        Returns:
            True if at least one entity is selected, False otherwise
        """
        try:
            self._validate_connection()
            doc = self._get_document("has_selection")

            # Use PickFirst selection set for reliable detection
            return doc.PickfirstSelectionSet.Count > 0

        except Exception as e:
            logger.debug(f"has_selection check failed: {e}")
            return False

    def get_selected_entity_handles(self) -> list[str]:
        """Get list of currently selected entity handles.

        Returns:
            List of entity handles (strings). Empty list if no selection.
        """
        try:
            self._validate_connection()
            doc = self._get_document("get_selected_entity_handles")

            handles = []

            # Use PickFirst selection set (most reliable)
            pickfirst = doc.PickfirstSelectionSet

            if pickfirst.Count > 0:
                for entity in pickfirst:
                    try:
                        handles.append(str(entity.Handle))
                    except Exception as e:
                        logger.debug(f"Failed to get handle for entity: {e}")
                        continue

                logger.info(f"Retrieved {len(handles)} selected entity handles")
                return handles

            logger.debug("No selected entities found")
            return []

        except Exception as e:
            logger.error(f"Failed to get selected entity handles: {e}")
            return []

    def get_selection_info(self) -> dict[str, Any]:
        """Get comprehensive information about current selection.

        Returns:
            Dictionary with:
            - count: Number of selected entities
            - handles: List of entity handles
            - types: List of entity ObjectNames
            - layers: Set of layers containing selected entities
        """
        try:
            self._validate_connection()
            doc = self._get_document("get_selection_info")

            info: dict[str, Any] = {
                "count": 0,
                "handles": [],
                "types": [],
                "layers": [],
            }

            pickfirst = doc.PickfirstSelectionSet
            info["count"] = pickfirst.Count

            if info["count"] > 0:
                layers_set: set[str] = set()

                for entity in pickfirst:
                    try:
                        info["handles"].append(str(entity.Handle))
                        info["types"].append(str(entity.ObjectName))
                        layers_set.add(str(entity.Layer))
                    except Exception as e:
                        logger.debug(f"Error extracting entity info: {e}")
                        continue

                info["layers"] = sorted(list(layers_set))

            return info

        except Exception as e:
            logger.error(f"Failed to get selection info: {e}")
            return {"count": 0, "handles": [], "types": [], "layers": []}

    # ========== Data Export ==========

    def _get_entities_to_process(
        self, document: Any, only_selected: bool = False
    ) -> list[Any]:
        """Get entities to process (all or selected).

        Optimized to use PickfirstSelectionSet for selected entities instead of
        iterating through entire ModelSpace.

        Args:
            document: AutoCAD document object
            only_selected: If True, get only selected entities. If False, get all.

        Returns:
            List of entity objects to process
        """
        entities_to_process = []

        if only_selected:
            # OPTIMIZED: Access selected entities directly from PickfirstSelectionSet
            # instead of iterating through entire ModelSpace looking for handles
            try:
                # Get the pickfirst selection set (current selection in AutoCAD)
                selection_set = document.PickfirstSelectionSet

                # Check if selection is empty
                if selection_set.Count == 0:
                    logger.info("No entities selected - returning empty list")
                    return []

                logger.info(f"Retrieving {selection_set.Count} selected entities")

                # Get entities directly from selection set (MUCH faster than iterating ModelSpace)
                for i in range(selection_set.Count):
                    try:
                        entities_to_process.append(selection_set.Item(i))
                    except Exception as e:
                        logger.debug(f"Failed to get selected entity at index {i}: {e}")
                        continue

            except Exception as e:
                logger.error(f"Failed to access PickfirstSelectionSet: {e}")
                logger.info("Falling back to handle-based selection method")

                # Fallback to old method if PickfirstSelectionSet fails
                selected_handles = self.get_selected_entity_handles()
                if not selected_handles:
                    logger.info("No entities selected - returning empty list")
                    return []

                logger.info(
                    f"Retrieving {len(selected_handles)} selected entities (fallback method)"
                )

                # Get entities by handle from ModelSpace
                try:
                    model_space = document.ModelSpace
                except Exception as e:
                    logger.error(f"Failed to access ModelSpace: {e}")
                    return []

                # Extract only selected entities
                selected_handles_set = set(selected_handles)

                for entity in model_space:
                    if str(entity.Handle) in selected_handles_set:
                        entities_to_process.append(entity)

        else:
            # Get all entities from ModelSpace
            try:
                model_space = document.ModelSpace
                entities_to_process = list(model_space)
            except Exception as e:
                logger.error(f"Failed to access ModelSpace: {e}")
                return []

        return entities_to_process

    def _extract_circle_properties(self, entity: Any) -> Dict[str, float]:
        """Extract Circle-specific geometry properties.

        Args:
            entity: Circle entity from AutoCAD

        Returns:
            Dictionary with radius, circumference, area, length
        """
        radius_val = self._fast_get_property(entity, "Radius")
        radius = float(radius_val) if radius_val else 0.0

        if radius > 0:
            circumference = 2 * math.pi * radius
            area = math.pi * radius * radius
        else:
            circumference = 0.0
            area = 0.0

        return {
            "Length": 0.0,
            "Area": round(area, 3) if area > 0 else 0.0,
            "Radius": round(radius, 3) if radius > 0 else 0.0,
            "Circumference": round(circumference, 3) if circumference > 0 else 0.0,
        }

    def _extract_arc_properties(self, entity: Any) -> Dict[str, float]:
        """Extract Arc-specific geometry properties.

        Args:
            entity: Arc entity from AutoCAD

        Returns:
            Dictionary with radius, length, circumference, area
        """
        radius_val = self._fast_get_property(entity, "Radius")
        length_val = self._fast_get_property(entity, "Length")

        radius = float(radius_val) if radius_val else 0.0
        length = float(length_val) if length_val else 0.0

        return {
            "Length": round(length, 3) if length > 0 else 0.0,
            "Area": 0.0,
            "Radius": round(radius, 3) if radius > 0 else 0.0,
            "Circumference": round(length, 3) if length > 0 else 0.0,  # Arc length
        }

    def _extract_line_properties(self, entity: Any) -> Dict[str, float]:
        """Extract Line-specific geometry properties.

        Args:
            entity: Line entity from AutoCAD

        Returns:
            Dictionary with length, area, radius, circumference
        """
        length_val = self._fast_get_property(entity, "Length")
        length = float(length_val) if length_val else 0.0

        return {
            "Length": round(length, 3) if length > 0 else 0.0,
            "Area": 0.0,
            "Radius": 0.0,
            "Circumference": 0.0,
        }

    def _extract_polyline_properties(self, entity: Any) -> Dict[str, float]:
        """Extract Polyline-specific geometry properties.

        Args:
            entity: Polyline entity from AutoCAD

        Returns:
            Dictionary with length, area, radius, circumference
        """
        length_val = self._fast_get_property(entity, "Length")
        area_val = self._fast_get_property(entity, "Area")

        length = float(length_val) if length_val else 0.0
        area = float(area_val) if area_val else 0.0

        return {
            "Length": round(length, 3) if length > 0 else 0.0,
            "Area": round(area, 3) if area > 0 else 0.0,
            "Radius": 0.0,
            "Circumference": 0.0,
        }

    def _extract_generic_properties(self, entity: Any) -> Dict[str, float]:
        """Extract generic entity properties (TEXT, DIMENSION, etc.).

        Args:
            entity: Generic entity from AutoCAD

        Returns:
            Dictionary with all geometry properties set to 0
        """
        return {
            "Length": 0.0,
            "Area": 0.0,
            "Radius": 0.0,
            "Circumference": 0.0,
        }

    def extract_drawing_data(self, only_selected: bool = False) -> list[dict]:
        """Extract drawing data (entities) with their properties.

        Optimized iteration through ModelSpace or selected entities with reduced COM calls.
        Uses property caching and batch processing for improved performance.

        Args:
            only_selected: If True, extract only selected entities. If False, extract all.
                          Defaults to False for backward compatibility.

        Returns:
            List of dictionaries with columns:
            - Handle: Entity handle (unique identifier)
            - ObjectType: Type of object (LINE, CIRCLE, LWPOLYLINE, etc.)
            - Layer: Layer name
            - Color: Color index (0-255) or color name
            - Length: Length (for linear objects)
            - Area: Area (for closed objects)
            - Radius: Radius (for circles and arcs)
            - Circumference: Circumference (2πr for circles, arc length for arcs)
            - Name: Name (for blocks, layers, etc.)
        """
        import time

        perf_start_total = time.perf_counter()

        try:
            self._validate_connection()
            document = self._get_document("extract_drawing_data")
            entities_data = []

            # Get entities to process (all or selected)
            perf_start_selection = time.perf_counter()
            entities_to_process = self._get_entities_to_process(document, only_selected)
            perf_selection_time = time.perf_counter() - perf_start_selection

            if not entities_to_process:
                logger.info("No entities to process - returning empty data")
                return []

            logger.info(
                f"[PERF] Entity selection/loading took {perf_selection_time:.3f}s ({len(entities_to_process)} entities)"
            )

            # Pre-build reverse color map for faster lookups
            color_map_reverse = {v: k for k, v in COLOR_MAP.items()}
            import math

            # Optimized iteration with reduced COM calls
            entity_count = 0
            error_count = 0
            perf_start_iteration = time.perf_counter()

            # Timing stats for property extraction
            perf_property_times: Dict[str, float] = {
                "basic": 0.0,  # Handle, ObjectName, Layer, Color, Name
                "geometry": 0.0,  # Length, Area
                "radius": 0.0,  # Radius, Circumference
            }

            # COM call statistics
            com_call_stats: Dict[str, Any] = {
                "total_calls": 0,
                "calls_by_type": {
                    "CIRCLE": 0,
                    "ARC": 0,
                    "LINE": 0,
                    "POLYLINE": 0,
                    "OTHER": 0,
                },
                "properties_skipped": 0,
            }

            try:
                # Progress tracking for large datasets
                total_entities = len(entities_to_process)
                progress_interval = 1000  # Log every 1000 entities
                sample_interval = 100  # Sample detailed timing every N entities

                for entity in entities_to_process:
                    entity_count += 1

                    # Progress logging for large datasets
                    if entity_count % progress_interval == 0:
                        elapsed = time.perf_counter() - perf_start_iteration
                        rate = entity_count / elapsed if elapsed > 0 else 0
                        logger.info(
                            f"[PERF] Progress: {entity_count}/{total_entities} entities "
                            f"({entity_count * 100 // total_entities}%) - {rate:.1f} entities/s"
                        )

                    try:
                        # Sample timing on a subset of entities
                        do_timing = entity_count % sample_interval == 0

                        # ========== PHASE 1: Basic Properties (ALWAYS) ==========
                        perf_t = time.perf_counter() if do_timing else 0.0

                        # CRITICAL: Only fetch absolutely necessary properties
                        handle = self._fast_get_property(entity, "Handle", "")
                        object_type = self._fast_get_property(
                            entity, "ObjectName", "Unknown"
                        )
                        layer = self._fast_get_property(entity, "Layer", "0")
                        calls_made = 3  # Handle, ObjectName, Layer

                        if do_timing:
                            perf_property_times["basic"] += time.perf_counter() - perf_t

                        # Pre-process object type for fast lookups
                        object_type_str = str(object_type)
                        object_type_upper = object_type_str.upper()

                        # OPTIMIZATION: Skip Color and Name for most entities (lazy fetch)
                        # Only fetch if entity type typically uses them
                        needs_color = True  # Most entities need color
                        needs_name = (
                            "BLOCK" in object_type_upper
                            or "INSERT" in object_type_upper
                        )

                        color = "ByLayer"
                        name = ""

                        if needs_color:
                            color_index = self._fast_get_property(entity, "Color", 256)
                            color = (
                                "ByLayer"
                                if color_index == 256
                                else color_map_reverse.get(
                                    color_index, str(color_index)
                                )
                            )
                            calls_made += 1
                        else:
                            com_call_stats["properties_skipped"] += 1

                        if needs_name:
                            name = self._fast_get_property(entity, "Name", "")
                            calls_made += 1
                        else:
                            com_call_stats["properties_skipped"] += 1

                        # ========== PHASE 2: Geometry Properties (SELECTIVE) ==========
                        length = 0.0
                        area = 0.0
                        radius = 0.0
                        circumference = 0.0

                        # OPTIMIZATION: Type-specific property extraction (minimize COM calls)
                        if "CIRCLE" in object_type_upper:
                            # Circle: Only need Radius (derive circumference + area from it)
                            perf_t = time.perf_counter() if do_timing else 0.0

                            radius_val = self._fast_get_property(entity, "Radius")
                            calls_made += 1  # Radius only (not Area, not Circumference)
                            com_call_stats["calls_by_type"]["CIRCLE"] += 1
                            com_call_stats["properties_skipped"] += (
                                2  # Skipped Area and Circumference COM calls
                            )

                            if radius_val is not None:
                                try:
                                    radius = float(radius_val)
                                    if radius > 0:
                                        circumference = 2 * math.pi * radius
                                        area = math.pi * radius * radius
                                except (ValueError, TypeError):
                                    pass

                            if do_timing:
                                perf_property_times["radius"] += (
                                    time.perf_counter() - perf_t
                                )

                        elif "ARC" in object_type_upper:
                            # Arc: Need Radius + Length (arc length = circumference)
                            perf_t = time.perf_counter() if do_timing else 0.0

                            radius_val = self._fast_get_property(entity, "Radius")
                            length_val = self._fast_get_property(entity, "Length")
                            calls_made += 2  # Radius + Length
                            com_call_stats["calls_by_type"]["ARC"] += 1
                            com_call_stats["properties_skipped"] += 1  # Skipped Area

                            if radius_val is not None:
                                try:
                                    radius = float(radius_val)
                                except (ValueError, TypeError):
                                    pass

                            if length_val is not None:
                                try:
                                    length = float(length_val)
                                    circumference = length  # Arc length
                                except (ValueError, TypeError):
                                    pass

                            if do_timing:
                                perf_property_times["radius"] += (
                                    time.perf_counter() - perf_t
                                )

                        elif (
                            "LINE" in object_type_upper
                            or "POLY" in object_type_upper
                            or "SPLINE" in object_type_upper
                        ):
                            # Linear entities: Only Length (skip Area unless it's a closed polyline)
                            perf_t = time.perf_counter() if do_timing else 0.0

                            length_val = self._fast_get_property(entity, "Length")
                            calls_made += 1  # Length only

                            if "LINE" in object_type_upper:
                                com_call_stats["calls_by_type"]["LINE"] += 1
                                com_call_stats["properties_skipped"] += (
                                    3  # Skipped Area, Radius, Circumference
                                )
                            elif "POLY" in object_type_upper:
                                com_call_stats["calls_by_type"]["POLYLINE"] += 1

                            if length_val is not None:
                                try:
                                    length = float(length_val)
                                except (ValueError, TypeError):
                                    pass

                            # LAZY: Only fetch Area if it's a polyline (might be closed)
                            if "POLY" in object_type_upper:
                                area_val = self._fast_get_property(entity, "Area")
                                calls_made += 1
                                com_call_stats["properties_skipped"] += (
                                    2  # Skipped Radius, Circumference
                                )
                                if area_val is not None:
                                    try:
                                        area = float(area_val)
                                    except (ValueError, TypeError):
                                        pass
                            else:
                                # Lines/Splines: skip Area entirely
                                com_call_stats["properties_skipped"] += 1

                            if do_timing:
                                perf_property_times["geometry"] += (
                                    time.perf_counter() - perf_t
                                )
                        else:
                            # Other entity types (TEXT, DIMENSION, etc.)
                            com_call_stats["calls_by_type"]["OTHER"] += 1
                            com_call_stats["properties_skipped"] += (
                                4  # All geometry properties
                            )

                        # Track total COM calls
                        com_call_stats["total_calls"] += calls_made

                        # ========== PHASE 3: Build Data Dictionary ==========
                        entity_data = {
                            "Handle": str(handle),
                            "ObjectType": object_type_str,
                            "Layer": str(layer),
                            "Color": color,
                            "Length": round(length, 3) if length > 0 else 0.0,
                            "Area": round(area, 3) if area > 0 else 0.0,
                            "Radius": round(radius, 3) if radius > 0 else 0.0,
                            "Circumference": (
                                round(circumference, 3) if circumference > 0 else 0.0
                            ),
                            "Name": str(name) if name else "",
                        }
                        entities_data.append(entity_data)

                    except Exception as e:
                        logger.debug(
                            f"Failed to extract entity data (entity #{entity_count}): {e}"
                        )
                        error_count += 1
                        continue

            except Exception as e:
                logger.error(f"Failed to iterate selected entities: {e}")
                return []

            perf_iteration_time = time.perf_counter() - perf_start_iteration
            perf_total_time = time.perf_counter() - perf_start_total

            logger.info(
                f"Extracted data from {len(entities_data)} entities "
                f"(processed {entity_count}, {error_count} errors)"
            )
            logger.info(
                f"[PERF] Entity iteration/extraction took {perf_iteration_time:.3f}s ({entity_count / perf_iteration_time:.1f} entities/s)"
            )

            # Detailed property timing breakdown
            samples_count = entity_count // sample_interval
            if samples_count > 0:
                avg_basic = (perf_property_times["basic"] / samples_count) * 1000
                avg_geometry = (perf_property_times["geometry"] / samples_count) * 1000
                avg_radius = (perf_property_times["radius"] / samples_count) * 1000
                logger.info(
                    f"[PERF] Property extraction (avg per entity): "
                    f"basic={avg_basic:.2f}ms, geometry={avg_geometry:.2f}ms, radius={avg_radius:.2f}ms"
                )

            # COM call optimization statistics
            total_calls = com_call_stats["total_calls"]
            skipped_calls = com_call_stats["properties_skipped"]
            potential_calls = total_calls + skipped_calls
            savings_pct = (
                (skipped_calls / potential_calls * 100) if potential_calls > 0 else 0
            )

            logger.info(
                f"[PERF] COM calls: {total_calls:,} made, {skipped_calls:,} skipped "
                f"({savings_pct:.1f}% reduction)"
            )
            logger.info(
                f"[PERF] Entity type breakdown: "
                f"CIRCLE={com_call_stats['calls_by_type']['CIRCLE']}, "
                f"ARC={com_call_stats['calls_by_type']['ARC']}, "
                f"LINE={com_call_stats['calls_by_type']['LINE']}, "
                f"POLY={com_call_stats['calls_by_type']['POLYLINE']}, "
                f"OTHER={com_call_stats['calls_by_type']['OTHER']}"
            )

            logger.info(f"[PERF] Total extraction time: {perf_total_time:.3f}s")
            return entities_data

        except Exception as e:
            logger.error(f"Failed to extract drawing data: {e}")
            return []

    def _safe_get_property(
        self, obj: Any, property_name: str, default: Any = None
    ) -> Any:
        """Safely get a COM object property with fallback value.

        Args:
            obj: COM object
            property_name: Name of property to get
            default: Default value if property access fails

        Returns:
            Property value or default
        """
        try:
            return getattr(obj, property_name)
        except Exception as e:
            logger.debug(f"Failed to get property {property_name}: {e}")
            return default

    def _fast_get_property(
        self, obj: Any, property_name: str, default: Any = None
    ) -> Any:
        """Fast version of _safe_get_property without logging for bulk operations.

        Use this in tight loops where logging overhead is significant.

        Args:
            obj: COM object
            property_name: Name of property to get
            default: Default value if property access fails

        Returns:
            Property value or default
        """
        try:
            return getattr(obj, property_name)
        except Exception:
            return default

    def export_to_excel(self, filepath: str = "drawing_data.xlsx") -> bool:
        """Export drawing data to Excel file.

        Uses the configured output directory from config.json for security,
        similar to save_drawing(). If only filename provided, saves to output directory.

        Args:
            filepath: Path to output Excel file (default: "drawing_data.xlsx")
                     - If filename only, saved to config output directory
                     - If path provided, must be within output directory

        Returns:
            True if successful, False otherwise
        """
        import time

        perf_start_total = time.perf_counter()

        try:
            from pathlib import Path
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            perf_start_setup = time.perf_counter()
            config = get_config()

            # SECURITY: Resolve output directory first (reference for validation)
            output_dir = Path(config.output.directory).expanduser().resolve()

            # ========== Determine Directory (filepath) ==========
            if filepath:
                # If filepath provided, extract directory part
                dir_part = str(Path(filepath).parent)
                if dir_part and dir_part != ".":
                    export_dir = dir_part
                else:
                    export_dir = str(output_dir)
            else:
                export_dir = str(output_dir)

            # Convert to absolute path (required by security validation)
            export_dir_path = Path(export_dir).expanduser().resolve()

            # SECURITY: Verify the directory is within the configured output directory
            try:
                export_dir_path.relative_to(output_dir)
            except ValueError:
                logger.error(
                    f"Security: Attempted to export outside output directory. "
                    f"Requested: {export_dir_path}, Allowed: {output_dir}"
                )
                return False

            # Create directory if it doesn't exist
            export_dir_path.mkdir(parents=True, exist_ok=True)

            # Get filename and construct full path
            filename = Path(filepath).name if filepath else "drawing_data.xlsx"
            full_filepath = export_dir_path / filename

            perf_setup_time = time.perf_counter() - perf_start_setup
            logger.info(f"[PERF] Export setup took {perf_setup_time:.3f}s")

            # Extract data
            perf_start_extract = time.perf_counter()
            data = self.extract_drawing_data()
            perf_extract_time = time.perf_counter() - perf_start_extract
            logger.info(f"[PERF] Data extraction took {perf_extract_time:.3f}s")

            if not data:
                logger.warning("No data to export")
                return False

            # Create workbook
            perf_start_workbook = time.perf_counter()
            workbook: Any = Workbook()
            worksheet: Any = workbook.active
            if worksheet is None:
                logger.error("Failed to create worksheet")
                return False

            worksheet.title = "Drawing Data"

            # Define columns
            columns = [
                "Handle",
                "ObjectType",
                "Layer",
                "Color",
                "Length",
                "Area",
                "Radius",
                "Circumference",
                "Name",
            ]

            # Write headers with styling
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")

            for col_idx, column_name in enumerate(columns, 1):
                cell: Any = worksheet.cell(row=1, column=col_idx)
                if cell is not None:
                    cell.value = column_name
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            perf_workbook_time = time.perf_counter() - perf_start_workbook
            logger.info(
                f"[PERF] Workbook creation and headers took {perf_workbook_time:.3f}s"
            )

            # Write data
            perf_start_write = time.perf_counter()
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, column_name in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell is not None:
                        value = row_data.get(column_name)
                        cell.value = value
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                        # Apply number format for numeric columns (3 decimals)
                        if column_name in [
                            "Length",
                            "Area",
                            "Radius",
                            "Circumference",
                        ] and isinstance(value, (int, float)):
                            cell.number_format = (
                                "0.000"  # Excel format: always 3 decimals
                            )

            perf_write_time = time.perf_counter() - perf_start_write
            logger.info(
                f"[PERF] Writing {len(data)} rows of data took {perf_write_time:.3f}s ({len(data) / perf_write_time:.1f} rows/s)"
            )

            # Auto-adjust column widths
            perf_start_autofit = time.perf_counter()
            for col_idx, column_name in enumerate(columns, 1):
                max_length = len(column_name)
                for row_idx in range(2, len(data) + 2):
                    cell_obj: Any = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = (
                        str(cell_obj.value or "") if cell_obj is not None else ""
                    )
                    max_length = max(max_length, len(cell_value))
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

            perf_autofit_time = time.perf_counter() - perf_start_autofit
            logger.info(
                f"[PERF] Auto-fitting column widths took {perf_autofit_time:.3f}s"
            )

            # Freeze first row (header) so it remains visible when scrolling
            worksheet.freeze_panes = "A2"

            # ========== Create Layers Sheet ==========
            perf_start_layers = time.perf_counter()
            # OPTIMIZED: Pass extracted data to avoid re-iterating ModelSpace
            layers_info = self.get_layers_info(entity_data=data)
            layers_sheet: Any = workbook.create_sheet("Layers")
            if layers_sheet is not None:
                # Define columns for layers sheet
                layer_columns = [
                    "Name",
                    "ObjectCount",
                    "Color",
                    "Linetype",
                    "Lineweight",
                    "IsLocked",
                    "IsVisible",
                ]

                # Write headers with styling
                for col_idx, column_name in enumerate(layer_columns, 1):
                    header_cell: Any = layers_sheet.cell(row=1, column=col_idx)
                    if header_cell is not None:
                        header_cell.value = column_name
                        header_cell.fill = header_fill
                        header_cell.font = header_font
                        header_cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

                # Write layer data
                for row_idx, layer_data in enumerate(layers_info, 2):
                    for col_idx, column_name in enumerate(layer_columns, 1):
                        data_cell: Any = layers_sheet.cell(row=row_idx, column=col_idx)
                        if data_cell is not None:
                            value = layer_data.get(column_name)
                            data_cell.value = value
                            # Center align for boolean and count columns
                            if column_name in ["ObjectCount", "IsLocked", "IsVisible"]:
                                data_cell.alignment = Alignment(
                                    horizontal="center", vertical="center"
                                )
                            else:
                                data_cell.alignment = Alignment(
                                    horizontal="left", vertical="center"
                                )

                # Auto-adjust column widths for layers sheet
                for col_idx, column_name in enumerate(layer_columns, 1):
                    max_length = len(column_name)
                    for row_idx in range(2, len(layers_info) + 2):
                        width_cell: Any = layers_sheet.cell(row=row_idx, column=col_idx)
                        cell_value = (
                            str(width_cell.value or "")
                            if width_cell is not None
                            else ""
                        )
                        max_length = max(max_length, len(cell_value))
                    col_letter = get_column_letter(col_idx)
                    layers_sheet.column_dimensions[col_letter].width = min(
                        max_length + 2, 50
                    )

                # Freeze first row (header) so it remains visible when scrolling
                layers_sheet.freeze_panes = "A2"

            perf_layers_time = time.perf_counter() - perf_start_layers
            logger.info(f"[PERF] Creating layers sheet took {perf_layers_time:.3f}s")

            # Save workbook
            perf_start_save = time.perf_counter()
            workbook.save(str(full_filepath))
            perf_save_time = time.perf_counter() - perf_start_save
            perf_total_time = time.perf_counter() - perf_start_total

            logger.info(f"[PERF] Saving workbook took {perf_save_time:.3f}s")
            logger.info(f"[PERF] Total export time: {perf_total_time:.3f}s")
            logger.info(
                f"Exported {len(data)} entities and {len(layers_info)} layers to {full_filepath}"
            )
            return True

        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}")
            return False

    # ========== Undo/Redo ==========

    def undo(self, count: int = 1) -> bool:
        """Undo last action(s).

        Args:
            count: Number of operations to undo (default: 1)

        Returns:
            True if successful, False otherwise
        """
        try:
            self._validate_connection()
            if count < 1:
                logger.warning(f"Invalid undo count: {count}. Must be >= 1")
                return False

            app = self._get_application("undo")
            app.ActiveDocument.SendCommand(f"_undo {count}\n")
            logger.info(f"Undo executed ({count} operation(s))")
            return True
        except Exception as e:
            logger.error(f"Failed to undo: {e}")
            return False

    def redo(self, count: int = 1) -> bool:
        """Redo last undone action(s).

        Args:
            count: Number of operations to redo (default: 1)

        Returns:
            True if successful, False otherwise
        """
        try:
            self._validate_connection()
            if count < 1:
                logger.warning(f"Invalid redo count: {count}. Must be >= 1")
                return False

            app = self._get_application("redo")
            app.ActiveDocument.SendCommand(f"_redo {count}\n")
            logger.info(f"Redo executed ({count} operation(s))")
            return True
        except Exception as e:
            logger.error(f"Failed to redo: {e}")
            return False

    # ========== Helper Methods ==========

    def _validate_connection(self) -> None:
        """Raise error if not connected."""
        if not self.is_connected():
            raise CADOperationError("connection", "Not connected to CAD application")
        if self.document is None:
            raise CADOperationError("connection", "Document is not available")

    def _get_document(self, operation: str = "operation") -> Any:
        """Get document with validation. Raises if not available."""
        self._validate_connection()
        if self.document is None:
            raise CADOperationError(operation, "Document not available")
        return self.document

    def _get_application(self, operation: str = "operation") -> Any:
        """Get application with validation. Raises if not available."""
        if self.application is None:
            raise CADOperationError(operation, "Application not available")
        return self.application

    def _wait_for(
        self,
        condition: Callable[[], bool],
        timeout: float = 20.0,
        interval: float = 0.1,
    ) -> bool:
        """Wait for a condition with timeout (replaces brittle time.sleep).

        Args:
            condition: Callable that returns True when condition is met
            timeout: Maximum seconds to wait (default: 20.0)
            interval: Check interval in seconds (default: 0.1)

        Returns:
            True if condition met before timeout, False otherwise
        """
        start_time = time.time()
        while time.time() - start_time < timeout:
            try:
                if condition():
                    return True
            except Exception:
                pass
            time.sleep(interval)
        return False

    def _delete_selection_set(self, document: Any, name: str) -> None:
        """Delete selection set if it exists (helper to reduce repetition)."""
        try:
            document.SelectionSets.Item(name).Delete()
        except Exception:
            pass

    def _to_variant_array(self, point: Point):
        """Convert 3D point to COM variant array."""
        return win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            [float(point[0]), float(point[1]), float(point[2])],
        )

    def _points_to_variant_array(self, points: List[Point]):
        """Convert list of 3D points to COM variant array (flattened)."""
        flat_array = []
        for point in points:
            flat_array.extend([float(point[0]), float(point[1]), float(point[2])])

        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, flat_array)

    def _objects_to_variant_array(self, objects: List[Any]) -> Any:
        """Convert list of COM objects to variant array for CopyObjects.

        Args:
            objects: List of COM entity objects

        Returns:
            VARIANT array of COM objects for CopyObjects method
        """
        return win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, objects
        )

    def _int_array_to_variant(self, values: tuple | list) -> Any:
        """Convert list of integers to COM variant array (for DXF filter codes)."""
        return win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_I2, [int(v) for v in values]
        )

    def _mixed_array_to_variant(self, values: tuple | list) -> Any:
        """Convert list of mixed types to COM variant array (for DXF filter data)."""
        variant_list: List[Any] = []
        for val in values:
            if isinstance(val, str):
                variant_list.append(val)
            elif isinstance(val, (int, float)):
                variant_list.append(val)
            else:
                variant_list.append(str(val))

        return win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, variant_list
        )

    def _to_radians(self, degrees: float) -> float:
        """Convert degrees to radians."""
        return degrees * math.pi / 180.0

    def _get_color_index(self, color_name: str) -> int:
        """Get CAD color index from color name."""
        color_name = color_name.lower().replace(" ", "_")
        return COLOR_MAP.get(color_name, 7)  # Default white

    def _apply_properties(
        self,
        entity: Any,
        layer: str,
        color: str | int,
        lineweight: int = 0,
    ) -> None:
        """Apply common properties to an entity."""
        try:
            entity.Layer = layer
            if isinstance(color, str):
                color = self._get_color_index(color)
            entity.Color = color
            if lineweight > 0:
                entity.LineWeight = self.validate_lineweight(lineweight)
        except Exception as e:
            logger.warning(f"Failed to apply properties: {e}")

    def _track_entity(self, entity: Any, entity_type: str) -> None:
        """Track entity in drawing state."""
        try:
            self._drawing_state["entities"].append(
                {
                    "handle": str(entity.Handle),
                    "type": entity_type,
                    "object_name": entity.ObjectName,
                }
            )
        except Exception as e:
            logger.warning(f"Failed to track entity: {e}")
